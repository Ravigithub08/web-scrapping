import numpy as np
from bs4 import BeautifulSoup
import sqlite3
import pandas as pd
import openpyxl
import requests as req
import re
import pymysql as mysql

#connect data base
db=mysql.connect(
    host="localhost",
    user="root",
    password="",
    port=3308,
    database="pydb"
)

cursor=db.cursor()
qry='CREATE TABLE IF NOT EXISTS WEB (col1 VARCHAR(255),col2 VARCHAR(255),col3 VARCHAR(255),col4 VARCHAR(255),col5 VARCHAR(255),col6 VARCHAR(255),col7 VARCHAR(255),col8 VARCHAR(255),col9 VARCHAR(255),col10 VARCHAR(255),col11 VARCHAR(255),col12 VARCHAR(255))'
cursor.execute(qry)


excel=openpyxl.Workbook()
sheet=excel.active
sheet.title="boolist"

request=req.get("https://en.wikibooks.org/wiki/Wikibooks_Stacks/Departments")           #get data from specified website

main_data=BeautifulSoup(request.text,'html.parser')
book_data=main_data.find('div',class_="vector-body").find_all("table")

data=[]
d={}
val=[]
for i in range(len(book_data)):
    if i==0:                                                #we don't need data from first table
        pass
    else:
        col=i
        topics=book_data[i].find("td").div.a.text
        
        sheet.cell(row=1,column=col,value=topics)           #set data to excel work book in specified cell
        if i !=3 and i!=11:
            td=book_data[i].find("td").find_all('ul')

            head=td[1].find_all('li')                       #table 1 and 11 only having data in first td rest of td having data in second td tag
        else:
            td=book_data[i].find("td").find_all('ul')

            head=td[0].find_all('li')
        for i in range(len(head)):
            
            y=head[i].a.text
            ind=i+2
            txt=re.sub("Shelf:","",y)
            sheet.cell(row=ind,column=col,value=txt)        #set data to excel work book in specified cell
            
            val.append(txt)
            d[topics]=val
            
        val=[]


max_len=max(len(v) for v in d.values())
mod={k:v+[None]*(max_len-len(v)) for k,v in d.items()}   #check all rows must be same lenth if not set none value 

df=pd.DataFrame.from_dict(mod)                              #convert data to pandas data frame to table look
t=[]

for j in range(len(df)):                                   #send data to mysql database using for loop
    loc=df.iloc[j]
    for i in range(len(loc)):
    
        t.append((loc.iloc[i]))
    cursor.execute("INSERT INTO web (col1,col2,col3,col4,col5,col6,col7,col8,col9,col10,col11,col12) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",t)
    t=[]

db.commit()
cursor.close()
db.close()
excel.save("book.xlsx")
print(df)


    